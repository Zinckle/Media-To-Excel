from time import sleep

import openpyxl as openpyxl
from PIL import Image
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import webbrowser
import os

if __name__ == '__main__':
    # imageName = input("Please enter an image name:")
    test = False
    images = os.listdir('Images')

    wb = openpyxl.Workbook()
    counter = 0

    for imageName in images:
        print(imageName)

        wb.create_sheet(str(counter))
        wb.active = wb[str(counter)]
        ws = wb.active
        ws.title = str(counter)

        counter += 1

        with Image.open('Images/' + imageName) as inputImage:
            # inputImage.show()
            rgbInputImage = inputImage.convert('RGB')
            width = inputImage.width
            height = inputImage.height

        for hPixel in range(1, height):
            shift = hPixel - 1 if hPixel > 1 else hPixel
            index = len(ws['A']) - shift
            for wPixel in range(1, width):
                r, g, b = rgbInputImage.getpixel((wPixel, hPixel))
                columnLetter = get_column_letter(wPixel)

                rVal = 'FF%02x%02x%02x' % (r, 0, 0)
                gVal = 'FF%02x%02x%02x' % (0, g, 0)
                bVal = 'FF%02x%02x%02x' % (0, 0, b)

                # ws[columnLetter + str(index + hPixel)] = r
                ws[columnLetter + str(index + hPixel)].fill = PatternFill(start_color=rVal, end_color=rVal,
                                                                          fill_type="solid")
                # ws[columnLetter + str(index + hPixel + 1)] = g
                ws[columnLetter + str(index + hPixel + 1)].fill = PatternFill(start_color=gVal, end_color=gVal,
                                                                              fill_type="solid")
                # ws[columnLetter + str(index + hPixel + 2)] = b
                ws[columnLetter + str(index + hPixel + 2)].fill = PatternFill(start_color=bVal, end_color=bVal,
                                                                              fill_type="solid")
        ws.sheet_view.zoomScale = 5
        ws.sheet_view.showGridLines = False
        # if test:
        # os.system("taskkill /im EXCEL.EXE /f")
        # test = True
    wb.remove(wb['Sheet'])
    wb.save(filename='sample_book.xlsx')
    webbrowser.open('sample_book.xlsx')
    # sleep(1)
    # os.system("taskkill /im EXCEL.EXE /f")
